import os
import tempfile
from docx import Document
from openpyxl import load_workbook

from converters.word_reader import extract_tables_from_word
from converters.excel_writer import write_single_table_to_excel


def test_full_pipeline():
    with tempfile.TemporaryDirectory() as tmp_dir:

        # 创建 Word
        word_path = os.path.join(tmp_dir, "test.docx")
        doc = Document()
        table = doc.add_table(rows=2, cols=2)

        table.cell(0, 0).text = "Name"
        table.cell(0, 1).text = "Age"
        table.cell(1, 0).text = "Alice"
        table.cell(1, 1).text = "18"

        doc.save(word_path)

        # 提取表格
        tables = extract_tables_from_word(word_path)

        excel_path = os.path.join(tmp_dir, "out.xlsx")

        write_single_table_to_excel(
            tables[0]["table_data"],
            excel_path
        )

        # 验证 Excel
        wb = load_workbook(excel_path)
        ws = wb.active

        assert ws.cell(1, 1).value == "Alice"
        assert ws.cell(1, 2).value == 18