import os
import tempfile
import pytest
from openpyxl import load_workbook

from converters.excel_writer import convert_value, write_single_table_to_excel


# =========================
# 测试 convert_value
# =========================

def test_convert_value_int():
    assert convert_value("123") == 123


def test_convert_value_float():
    assert convert_value("123.45") == 123.45


def test_convert_value_string():
    assert convert_value("abc") == "abc"


def test_convert_value_empty_string():
    assert convert_value("") == ""


def test_convert_value_none():
    assert convert_value(None) is None


def test_convert_value_with_spaces():
    assert convert_value("  42  ") == 42


def test_convert_value_invalid_number():
    assert convert_value("12abc") == "12abc"


# =========================
# 测试 write_single_table_to_excel
# =========================

def test_write_single_table_to_excel_basic():
    table_data = [
        ["1", "2", "3"],
        ["4.5", "text", ""]
    ]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        output_path = tmp.name

    try:
        write_single_table_to_excel(table_data, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        assert ws.cell(1, 1).value == 1
        assert ws.cell(1, 2).value == 2
        assert ws.cell(1, 3).value == 3

        assert ws.cell(2, 1).value == 4.5
        assert ws.cell(2, 2).value == "text"
        assert ws.cell(2, 3).value is None

    finally:
        os.remove(output_path)


def test_write_single_table_empty():
    table_data = []

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        output_path = tmp.name

    try:
        write_single_table_to_excel(table_data, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # 空表 -> 没有数据
        assert ws.max_row == 1
        assert ws.max_column == 1

    finally:
        os.remove(output_path)


def test_write_single_table_mixed_types():
    table_data = [
        ["001", "3.14", "hello", " "],
        [None, "100", "abc123", "0"]
    ]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        output_path = tmp.name

    try:
        write_single_table_to_excel(table_data, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        assert ws.cell(1, 1).value == 1         # "001" -> int
        assert ws.cell(1, 2).value == 3.14
        assert ws.cell(1, 3).value == "hello"
        assert ws.cell(1, 4).value is None        # " " -> None

        assert ws.cell(2, 1).value is None
        assert ws.cell(2, 2).value == 100
        assert ws.cell(2, 3).value == "abc123"
        assert ws.cell(2, 4).value == 0

    finally:
        os.remove(output_path)